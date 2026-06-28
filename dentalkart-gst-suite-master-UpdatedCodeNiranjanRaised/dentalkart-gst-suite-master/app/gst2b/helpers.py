import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


def find_header_row(df_raw, keywords=('gstin', 'gst no', 'party', 'supplier'), max_scan=10):
    for i in range(min(max_scan, len(df_raw))):
        row = df_raw.iloc[i]
        row_str = row.fillna('').astype(str).str.lower()
        if any(any(kw in cell for kw in keywords) for cell in row_str):
            return i
    return 0


def map_columns(df, column_map):
    actual_cols = list(df.columns)
    mapping = {}
    for std_name, candidates in column_map.items():
        found = None
        for candidate in candidates:
            cand_lower = candidate.lower()
            for col in actual_cols:
                if str(col).lower().strip() == cand_lower:
                    found = col
                    break
            if found:
                break
            for col in actual_cols:
                if cand_lower in str(col).lower():
                    found = col
                    break
            if found:
                break
        mapping[std_name] = found
    return mapping


def format_header(ws, row_num, color='366092'):
    for cell in ws[row_num]:
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            continue
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_col_width(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def load_excel_with_header(file_bytes, col_map, label):
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=15)
    header_row = find_header_row(df_raw)
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
    mapping = map_columns(df, col_map)
    return df, mapping, header_row
