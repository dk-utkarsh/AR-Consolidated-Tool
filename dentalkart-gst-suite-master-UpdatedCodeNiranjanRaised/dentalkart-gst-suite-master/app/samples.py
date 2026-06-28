"""Sample file generators for testing."""
import io
import openpyxl
from openpyxl.styles import PatternFill, Font


def _wb_headers(ws, headers, color='1F4E79'):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20


def generate_sample_sales():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sales Data"
    _wb_headers(ws, ['InvoiceNo','DateofInvoice','ExternalOrderNo','Order Channel',
        'Billing Name of Customer','State','ClientSKU','DescriptionofGoods','QTY','HSN Code',
        'GST No','E-invoice No.(IRN)','E-way bill number','Unit Price','TaxPercent',
        'Taxable Value','CGSTAmount','SGSTAmount','IGSTAmount','TaxAmount','OrderAmount',
        'Selling Price','Discount on item'])
    ws.append(['DK-INV-001','15-01-2026','ORD-1001','WhatsApp','Sample Customer','Delhi',
               'SKU001','Dental Mirror',10,'90184990','07AAACR1718R1ZP','IRN123456',
               'EWB789012',500,18,5000,450,450,0,900,5900,590,0])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def generate_sample_einvoice():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "b2b,sez,de"
    _wb_headers(ws, ['Invoice Number','Invoice Value','IRN','Invoice date',
                     'GSTIN/UIN of Recipient','Receiver Name'])
    ws.append(['DK-INV-001',5900,'IRN123456','15-01-2026','07AAACR1718R1ZP','Sample Customer'])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def generate_sample_ewaybill():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Outward Supply"
    ws.append([])
    _wb_headers(ws, ['Doc.No','Doc.Date','EWB No','EWB Date','Branch','Supply Type',
                     'Other Party GSTIN','TO GSTIN Info','Total Invoice Value','status'])
    ws.append(['DK-INV-001','15-01-2026','EWB789012','15-01-2026','Delhi',
               'Outward','07AAACR1718R1ZP','',5900,'Active'])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def generate_sample_creditnote():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Credit Notes"
    _wb_headers(ws, ['Return No','Return Date','InvoiceNo','Date of Invoice','Order Channel',
                     'Name of Customer','State','GST No','Return Amount'])
    ws.append(['RET-001','20-01-2026','DK-INV-001','15-01-2026','WhatsApp',
               'Sample Customer','Delhi','07AAACR1718R1ZP',2950])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def generate_sample_cn_einvoice():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "b2b-cr"
    ws.append([]); ws.append([])
    _wb_headers(ws, ['Unit','GSTIN/UIN of Recipient','Receiver Name','Note Number','Note Date',
                     'Note Type','Place of Supply','Reverse Charge','Note Supply Type','Note value',
                     'Applicable % of Tax Rate','Rate','Taxable Value','Integrated Tax',
                     'Central Tax','State/UT Tax','Cess Amount','IRN','IRN date','E-invoice status'])
    ws.append(['VIK','27AAACR1718R1ZP','Sample Customer','CN-001','15-01-2026','C',
               '07 - Delhi','N','Regular B2B',5900,None,18,5000,0,450,450,0,
               'IRN_CN_001','15-01-2026','Valid'])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()
